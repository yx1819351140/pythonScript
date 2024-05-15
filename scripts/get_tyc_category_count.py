# -*- coding:UTF-8 -*-
# @Time    : 2024/5/13 16:39
# @Author  : yangxin
# @Email   : yang1819351140@163.com
# @IDE     : PyCharm
# @File    : get_tyc_category_count.py
# @Project : pythonScript
# @Software: PyCharm
"""
获取指定表格中对应分类的天眼查公司数量
"""
import time

import requests
from lxml import etree
from openpyxl import load_workbook


def get_company_count(category_name='商务服务'):
    try:
        headers = {
            'Host': 'capi.tianyancha.com',
            'Connection': 'keep-alive',
            'Content-Length': '77',
            'Content-Type': 'application/json',
            'xweb_xhr': '1',
            'Authorization': '0###oo34J0XMNshkIjHzjrvV1tS5pLFA###1715670236959###ac9e661088c6a54db49610ad186912d3',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 MicroMessenger/7.0.20.1781(0x6700143B) NetType/WIFI MiniProgramEnv/Windows WindowsWechat/WMPF WindowsWechat(0x63090a13) XWEB/9129',
            'version': 'TYC-XCX-WX',
            'Accept': '*/*',
            'Sec-Fetch-Site': 'cross-site',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Dest': 'empty',
            'Referer': 'https://servicewechat.com/wx9f2867fc22873452/105/page-frame.html',
            'Accept-Encoding': 'gzip, deflate, br',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }
        url = 'https://capi.tianyancha.com/cloud-tempest/app/searchCompany'
        data = {"sortType": 0, "pageSize": 20, "pageNum": 1, "word": category_name, "allowModifyQuery": 1}
        res = requests.post(url, headers=headers, json=data)
        # print(res.text)
        company_count = res.json().get('data', {}).get('companyTotalStr')
        # print(company_count)
        if not company_count:
            return '0'
        return company_count
    except Exception as e:
        # print(e)
        return '0'


def run():
    workbook = load_workbook(filename='data/contrast_20240508.xlsx')
    sheet = workbook.active
    err_count = 0
    category_list = []
    with open('data/contrast_20240508.txt', 'r') as f:
        i = int(f.readlines()[-9].split('   ')[0])
    while i < sheet.max_row + 1:
        category = sheet[f'B{i}'].value
        if category in category_list:
            count = sheet[f'H{i - 1}'].value
            sheet[f'H{i}'] = count
        else:
            count = get_company_count(category)
            sheet[f'H{i}'] = count
            category_list.append(category)
        sub_category = sheet[f'D{i}'].value
        count1 = get_company_count(category + sub_category)
        sheet[f'I{i}'] = count1
        if count1 != '0':
            err_count = 0
        else:
            err_count += 1
        if err_count >= 10:
            time.sleep(3)
            with open('data/contrast_20240508.txt', 'r') as f:
                i = int(f.readlines()[-9].split('   ')[0])
        else:
            backup_data = f'{i}   {category}   {sub_category}   {count}   {count1}'
            with open('data/contrast_20240508.txt', 'a') as f:
                f.write(backup_data + '\n')
            print(backup_data)
            i += 1
    workbook.save('data/contrast_20240508.xlsx')


if __name__ == '__main__':
    # get_company_count('稀土镧')
    run()
